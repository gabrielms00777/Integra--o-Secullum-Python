import os
from openpyxl import load_workbook
from api_client import APIConfig

class Funcoes:
    def __init__(self, client: APIConfig):
        self.client = client

    def all(self):
        return self.client.get('Funcoes')
    
    def create(self, description):
        data = {
            'Descricao': description,
        }
        return self.client.post('Funcoes', data)
    
    def delete(self, description):
        endpoint = f"Funcoes?descricao={description}"
        return self.client.delete(endpoint)
    
    def import_functions(self, filename):
        self.client.verify_file_exists(filename)
        workbook = load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            fn_name = row[0]
            if fn_name:
                response = self.create(fn_name)
                print(f"Função {fn_name} importado: {response}")
            
    def treat_justifications(self):
        action = input("Deseja (v)isualizar, (c)adastrar, (i)mportr, ou (d)eletar uma função? ").lower()
        match action:
            case 'v':
                print("Funções: ", self.all())
            case 'c':
                description = input("Informe o nome da função: ")
                response = self.create(description)
                print(f"Departamento {description} cadastrado: {response}")
            case 'd':
                description = input("Informe o nome da função que deseja deletar: ")
                response = self.delete(description)
                print(f"Departamento {description} deletado: {response}")
            case 'i':
                filename = input("Informe o caminho do arquivo que deseja importar: ")
                self.import_functions(filename)
            