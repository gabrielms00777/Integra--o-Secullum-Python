from openpyxl import load_workbook

workbook = load_workbook('departamentos.xlsm')
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
    dp_name = row[0]
    if dp_name:
        print(dp_name)