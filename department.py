import os
from openpyxl import load_workbook
from api_client import APIConfig

class Department:
    def __init__(self, client: APIConfig) -> None:
        self.client = client

    def all(self):
        return self.client.get('Departamentos')
    
    def create(self, description, n_folha=''):
        data = {
            'Descricao': description,
            'Nfolha': n_folha
        }
        return self.client.post('Departamentos', data=data)

    def delete(self, description):
        endpoint = f"Departamentos?descricao={description}"
        return self.client.delete(endpoint)
    
    def verify_file_exists(self, filename):
        if not os.path.isfile(filename):
            raise FileNotFoundError(f"O arquivo {filename} não foi encontrado no diretório.")

    def import_department(self, filename):
        self.verify_file_exists(filename)
        workbook = load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            dp_name = row[0]
            if dp_name:
                response = self.create(dp_name)
                print(f"Departamento {dp_name} importado: {response}")

    def treat_justifications(self):
        action = input("Deseja (v)isualizar, (c)adastrar, (i)mportr, ou (d)eletar um departamento? ").lower()
        match action:
            case 'v':
                departments = self.all()
                print("Departamentos:", departments)
            case 'c':
                description = input("Informe o nome do departamento: ")
                nfolha = input("Informe o número da folha (opcional): ")
                response = self.create(description, nfolha)
                print(f"Departamento {description} cadastrado: {response}")
            case 'd':
                description = input("Informe a descrição do departamento que deseja deletar: ")
                response = self.delete(description)
                print(f"Departamento {description} deletado: {response}")
            case 'i':
                filename = input("Informe o caminho do arquivo que deseja importar: ")
                self.import_department(filename)

